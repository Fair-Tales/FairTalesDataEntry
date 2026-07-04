#!/usr/bin/env python3
"""Standalone pilot-corpus import CLI for FairTalesDataEntry (issue #73).

Imports the PILOT study corpus (~200 primary-school picture books) into the
production storage stack — **Firestore** for the structured records/text and
**S3** (``sawimages``) for the page images — in exactly the same document shape
the Streamlit app writes today.

Sources (a sibling checkout of ``fair-tales-language-analysis``)
---------------------------------------------------------------
* ``Book-List-Final-NONA.xlsx`` (Sheet1) — one row per book: title, author,
  author gender (M/F), year of first publication, starting/ending story page,
  and protagonist/secondary-character summary columns.
* ``character_database.db`` (SQLite) — ``characters`` (name/gender/human/
  is_protagonist per book), ``protagonists`` and ``aliases``.
* ``text_pdfs/*.pdf`` — the scanned book pages (one PDF per book), most with an
  embedded text layer; image-only pages fall back to Claude vision OCR.
* ``data/book_dataframe.pickle`` — Title -> full book text, the human-validated
  text used for the published analysis. Loaded as the *cross-check reference*:
  each book's per-page extracted text is compared against it (word-overlap) and
  low-scoring books are flagged for review. ``data/book_dataframe.json`` is a
  byte-identical export kept only as a fallback if the pickle won't unpickle.

Per-page text pipeline (issue #73, plan "A")
--------------------------------------------
Two passes per book:

* **Pass 1 (extract):** render each page to JPEG (S3) and get its raw text — the
  PDF text layer where present, otherwise Claude vision OCR (Sonnet 5 by default,
  a detailed picture-book-aware prompt) for the image-only minority. Before
  paying for OCR, an image-only STORY page that is flanked on BOTH sides by
  text-layer story pages is put through a cheap text-only **neighbour-continuity
  judge** (``ai_continuity.check_narrative_continuity``, a Sonnet model): if the
  story reads continuously straight across it, the page is a genuine wordless
  spread and its OCR is SKIPPED (text stored empty, ``text_source=
  "skipped_wordless"``, verdict recorded for audit) — validated at ~81% of such
  OCR calls avoided with ~zero data loss (DECISIONS 010). Edges, runs of
  consecutive image-only pages, and any image-only neighbour always OCR; and a
  judge error always OCRs (never a false skip). Quote/tag/punctuation-only output
  is normalised to empty (a blank page). An OCR *call failure* is recorded
  distinctly from a genuine blank so it becomes a high-priority review flag
  instead of silently blanking real text (M1).
* **Pass 2 (clean + judge in context):** for each story page, one Claude call
  (a) strips extraction/OCR garbage AND print artefacts (page numbers, running
  heads, copyright boilerplate) **without** altering spelling or unusual/invented
  words (a token-subset guard discards any "clean" that adds or alters a word),
  and (b) judges the page on two axes given the neighbour context — does it read
  as sensible story text (``makes_sense``) and does it fit its neighbours
  (``fits_context``). The neighbour boundary is derived from POSITION, not text
  truthiness, so a run of missing pages is not each mistaken for the story's
  start/end (M4). A page failing either check is flagged ``needs_review``;
  failing **both**, or an empty page that breaks the narrative, marks it
  ``review_priority="high"``. A rejected clean and a failed clean/judge call are
  both flagged too. Flags are recorded on the page and aggregated onto the book.

The OCR and clean/judge calls use Anthropic **structured outputs** (guaranteed
valid JSON) and prompt caching on their static instruction prefix, and their
results are cached locally (``--cache-dir``) so a crash or ``--overwrite`` re-run
does not re-pay for them. ``MAX_CONSECUTIVE_AI_FAILURES`` consecutive AI failures
abort the run (safely resumable — book docs are written last).

Finally, the book's concatenated page text is cross-checked against the validated
pickle reference as a per-book confidence signal: the PRIMARY metric is
CONTAINMENT (recall of the validated words), with Jaccard kept as a secondary
"excess text" indicator.

Shared entities (``authors`` / ``illustrators`` / ``publishers``) are written
ONLY IF absent, so importing never clobbers a human-entered record (M2).

Design & safety
---------------
This script runs **outside Streamlit**, so it deliberately does NOT import the
Streamlit-coupled ``data_structures`` package. It loads ``.streamlit/secrets.toml``
directly and builds its own ``firestore.Client`` (project ``sawdataentry``),
``s3fs.S3FileSystem`` and ``anthropic.Anthropic`` clients, mirroring the app's
``FirestoreWrapper`` / uploader / ``utilities`` config, exactly like
``scripts/data_cleanup.py``. The Firestore document schema for every entity
(Book/Author/Illustrator/Publisher/Page/Character/Alias) is **replicated** here
by studying the corresponding ``data_structures/*.py`` classes — it is NOT
imported.

**DRY-RUN BY DEFAULT.** With no ``--execute`` flag the tool writes NOTHING to
Firestore or S3: it prints a per-book plan plus corpus totals, every
title-match gap / discrepancy, and a small sample of the AI illustrator/
publisher lookup and per-page text extraction, then exits. ``--execute``
performs the real import (idempotent — books already present in Firestore are
skipped). The live run must be performed by a maintainer where the real secrets
exist; it cannot connect from an isolated git worktree.

Usage
-----
    # Dry run (default — writes nothing):
    python scripts/import_pilot_data.py \
        --methods-dir ../fair-tales-language-analysis

    # Real import (writes to Firestore + S3 — run only intentionally):
    python scripts/import_pilot_data.py \
        --methods-dir ../fair-tales-language-analysis --execute

Mapping decisions (see scripts/PILOT_IMPORT.md for the full rationale)
----------------------------------------------------------------------
* Author gender:    M -> "Man",  F -> "Woman",  anything else -> "Unknown"
                    (AuthorForm.gender_options). "X and Y" co-authored rows with
                    a paired code ("M/F") take the FIRST person + FIRST code as
                    the author (the second name is usually the illustrator);
                    shared-surname forms ("Janet and Allen Ahlberg") and 3+-code
                    rows are flagged for manual entry instead (S2).
* Character gender: F -> "Female",  M -> "Male",  NGS -> "Non-specific"
                    (CharacterForm.gender_options; there is no "Non-binary"
                    option for characters, so unknowns fall back to
                    "Non-specific").
* Character human:  "H" -> True, "NH" -> False (whitespace stripped). The DB
                    also holds two stray values ("NO", "NGS") which are flagged
                    and default to the Character.human default (True).
* Character plural: no plural/singular indicator exists in the sheet or DB, so
                    ``plural`` is always False and flagged in the report.
* ethnicity/disability: not present in the pilot data -> "Not specified"
                    (the first CharacterForm option).
"""

from __future__ import annotations

import argparse
import hashlib
import inspect
import io
import json
import os
import re
import sqlite3
import sys
import tempfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Optional

# The reusable, Streamlit-free neighbour-continuity judge lives at the repo root
# (importable by both this standalone CLI and the live app's utilities.py). This
# script's own directory is ``scripts/``, so add the repo root to the path first.
_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)
from ai_continuity import check_narrative_continuity, should_skip_ocr  # noqa: E402

# ---------------------------------------------------------------------------
# Configuration.
# ---------------------------------------------------------------------------

#: S3 bucket holding book page images (matches the app / s3_constants.S3_BUCKET).
S3_BUCKET = "sawimages"

#: Firestore project id (matches the app's FirestoreWrapper / data_cleanup.py).
FIRESTORE_PROJECT = "sawdataentry"

#: Default path to the Streamlit secrets file (loaded directly, no Streamlit).
DEFAULT_SECRETS = ".streamlit/secrets.toml"

#: The user recorded as the author of every imported record.
ENTERED_BY = "pilot_import"

#: Claude model used for the illustrator/publisher/year web lookup. Defaults to
#: the most capable widely-released model for best accuracy (the corpus is
#: small, so cost is bounded and accuracy is prioritised per #73). Overridable
#: via --lookup-model.
DEFAULT_LOOKUP_MODEL = "claude-opus-4-8"

#: Claude model used for per-page OCR when a PDF page has no extractable text
#: layer. Defaults to Sonnet 5 — validated as equal-quality OCR at ~-60% cost vs
#: Opus on this corpus (DECISIONS 010). Overridable via --ocr-model.
DEFAULT_OCR_MODEL = "claude-sonnet-5"

#: Claude model used for the text-only neighbour-continuity judge that lets an
#: image-only story page flanked by text-layer pages SKIP OCR when the story
#: flows straight through (ai_continuity.check_narrative_continuity). The judge
#: was validated on a Sonnet model; overridable via --continuity-model.
DEFAULT_CONTINUITY_MODEL = "claude-sonnet-4-6"

#: Claude model used for the per-page clean-up + coherence/context judgement. The
#: garbage-strip is mechanical, but the makes_sense / fits_context judgement is a
#: reasoning task, so this defaults to a mid-tier model. Overridable via
#: --clean-model.
DEFAULT_CLEAN_MODEL = "claude-sonnet-4-6"

#: Per-book cross-check (secondary, "excess text" signal): symmetric Jaccard
#: word-overlap of extracted vs validated text. Books below this ratio are noted
#: but the primary flag uses containment (below). Confidence signal only.
DEFAULT_COMPARE_THRESHOLD = 0.60

#: Per-book cross-check (PRIMARY flag, #73 S3): asymmetric CONTAINMENT / recall =
#: fraction of the validated reference's words captured by the extracted text.
#: Unlike Jaccard it does not penalise OCR that legitimately reads MORE than the
#: sparse validated pickle. Books below this are flagged for manual review. This
#: is a confidence signal only; it changes no data.
DEFAULT_CONTAINMENT_THRESHOLD = 0.85

#: Circuit-breaker (#73 M1): this many CONSECUTIVE AI failures (OCR or
#: clean/judge) aborts the run with a clear message. The run is safely resumable
#: because a book's ``books`` document is written LAST, so an aborted book is
#: re-done on the next run and already-written books are skipped.
MAX_CONSECUTIVE_AI_FAILURES = 5

#: Default local result-cache directory (#73 S9). OCR and clean/judge results are
#: cached here keyed by content hash so a crash or ``--overwrite`` re-run does not
#: re-pay for the API calls. Overridable via ``--cache-dir``; ``--no-cache``
#: disables it. A corrupt/missing cache never crashes the run — it just misses.
DEFAULT_CACHE_DIR = os.path.join(tempfile.gettempdir(), "fairtales_pilot_import_cache")

#: Longest-edge pixel cap when rendering a PDF page to JPEG for S3.
DEFAULT_MAX_EDGE = 2000

#: JPEG quality for rendered page images.
DEFAULT_JPEG_QUALITY = 85

#: A PDF page whose extracted text-layer has at least this many non-whitespace
#: characters is treated as having a usable text layer; shorter/blank pages fall
#: back to AI OCR on --execute.
TEXT_LAYER_MIN_CHARS = 20

#: The character-gender options the app offers (text_content.CharacterForm).
#: Replicated here rather than imported (Streamlit-coupled). Kept in sync
#: manually; the mapper below only emits values from this list.
CHARACTER_GENDER_OPTIONS = ["Female", "Male", "Non-specific", "Transgender"]

#: The author-gender options the app offers (text_content.AuthorForm).
AUTHOR_GENDER_OPTIONS = ["Woman", "Man", "Non-binary", "Other", "Unknown"]

#: Default ethnicity / disability (first CharacterForm option); the pilot data
#: carries neither field.
DEFAULT_ETHNICITY = "Not specified"
DEFAULT_DISABILITY = "Not specified"


# ---------------------------------------------------------------------------
# Pure helpers (no I/O, unit-testable).
# ---------------------------------------------------------------------------

def normalise_title(title: object) -> str:
    """Normalise a book title for cross-source matching.

    Lower-cases, expands ``&`` to ``and``, unifies curly quotes, strips all
    non-alphanumeric runs to single spaces and trims. This collapses the
    Excel trailing-space / punctuation / case differences against the SQLite
    ``book`` values and the PDF filenames so the same book matches across all
    three sources.
    """
    if title is None:
        return ""
    text = str(title).strip().lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[‘’“”]", "'", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def book_document_id(title: str) -> str:
    """Book.document_id — ``title.lower().replace(" ", "_")`` (see book.py)."""
    return title.lower().replace(" ", "_")


def person_document_id(name: str) -> str:
    """Author/Illustrator/Publisher document_id — ``name.lower().replace(" ", "_")``."""
    return name.lower().replace(" ", "_")


def character_document_id(book_id: str, name: str) -> str:
    """Character.document_id — ``{book_id}_{name}`` (see character.py)."""
    return f"{book_id}_{name.replace(' ', '_').lower()}"


def alias_document_id(book_id: str, alias_name: str) -> str:
    """Alias.document_id — ``{book_id}_{alias}`` (see alias.py)."""
    return f"{book_id}_{alias_name.replace(' ', '_').lower()}"


def page_document_id(book_id: str, page_number: int) -> str:
    """Page.document_id — ``{book_id}_{page_number}`` (see page.py)."""
    return f"{book_id}_{page_number}"


def s3_page_path(title: str, page_number: int) -> str:
    """S3 object path for a page image — ``sawimages/{title}/page_N.jpg``.

    Mirrors pages/uploader.py, which sets ``photos_url = f"sawimages/{title}"``
    (the raw title, not the underscored document id) and writes each page to
    ``f"{photos_url}/page_{n}.jpg"``.
    """
    return f"{S3_BUCKET}/{title}/page_{page_number}.jpg"


def book_photos_url(title: str) -> str:
    """The ``photos_url`` the app stores for a book — ``sawimages/{title}``."""
    return f"{S3_BUCKET}/{title}"


def split_name(full_name: str) -> tuple[str, str]:
    """Split a full author name into (forename, surname).

    Last whitespace token is the surname; everything before it is the forename.
    Multi-author strings (e.g. "Michael Rosen and Helen Oxenbury") are kept
    whole with the final word as the surname — best effort; such rows are
    flagged in the report.
    """
    parts = str(full_name).strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return " ".join(parts[:-1]), parts[-1]


def map_author_gender(value: object) -> str:
    """Map the sheet's author-gender code to an AuthorForm option.

    ``M`` -> "Man", ``F`` -> "Woman", anything else (blank, "M/F" multi-author,
    unknown) -> "Unknown".
    """
    code = str(value or "").strip().upper()
    if code == "M":
        return "Man"
    if code == "F":
        return "Woman"
    return "Unknown"


def is_unknown(value: object) -> bool:
    """Case-insensitive "no confirmed value" test for AI lookup results (#73 S1).

    Treats ``""``, ``"unknown"`` (any case), ``"n/a"``, ``"none"`` and ``"null"``
    as unknown, so a lowercase ``"unknown"`` reply never becomes a doc named
    ``unknown``.
    """
    return str(value or "").strip().lower() in ("", "unknown", "n/a", "none", "null")


@dataclass(frozen=True)
class AuthorParse:
    """Result of parsing a (possibly multi-author) sheet row (#73 S2).

    ``author_name`` / ``gender_code`` are what to import as the book's author
    (``author_name`` empty means "import no author"); ``second_name`` is the
    other credited person for a two-author row (usually the illustrator);
    ``needs_review`` / ``note`` flag rows a human must finish by hand.
    """

    author_name: str
    gender_code: str
    second_name: str
    needs_review: bool
    note: str


def parse_author_field(author: object, gender_code: object) -> AuthorParse:
    """Recover the author (and gender) from a "X and Y" sheet row (#73 S2).

    ~25% of the corpus lists two people ("Michael Rosen and Helen Oxenbury")
    with a paired gender code ("M/F"). Naively this discards the key gender
    variable and invents a merged author. Rules:

    * Exactly one " and " AND a two-single-letter code ("M/F"): take the FIRST
      person as the author with the FIRST code, and return the second name
      separately (usually the illustrator).
    * IMPORTANT shared-surname edge ("Janet and Allen Ahlberg", "Ronda and David
      Armitage"): the first part is a bare forename (single token, no surname),
      so importing "Janet" as an author would be wrong — flag the book for
      manual author entry instead of importing.
    * A code with 3+ parts ("M/F/M") → keep the row's author string but set the
      gender to Unknown and flag it.
    * Genuine single-author rows are returned unchanged.
    """
    author = str(author or "").strip()
    code = str(gender_code or "").strip()
    parts = author.split(" and ")

    if "/" in code:
        codes = [c.strip() for c in code.split("/")]
        # 3+ gender codes (e.g. "M/F/M", the "Oi Dog!" case) — too ambiguous to
        # split; keep the row's author string but set gender Unknown and flag it.
        if len(codes) >= 3:
            return AuthorParse(
                author, "", "", True,
                "multi-author with 3+ gender codes; author gender set to Unknown",
            )
        # Exactly one " and " AND two single-letter codes -> a splittable pair.
        if len(parts) == 2 and len(codes) == 2 and all(
            len(c) == 1 and c.isalpha() for c in codes
        ):
            first_person = parts[0].strip()
            second_person = parts[1].strip()
            # Shared-surname bare-forename form ("Janet and Allen Ahlberg"): the
            # first part has no surname, so importing it as an author is wrong.
            if len(first_person.split()) < 2:
                return AuthorParse(
                    "", "", second_person, True,
                    "multi-author with shared surname "
                    "('Forename and Forename Surname'); enter the author manually",
                )
            return AuthorParse(first_person, codes[0], second_person, False, "")

    # Genuine single-author (or unusual) row — behaviour unchanged.
    return AuthorParse(author, code, "", False, "")


def _as_bool(value: object, default: bool = True) -> bool:
    """Parse a judge boolean, correctly handling the strings 'true'/'false' (#73 S6).

    ``bool("false")`` is ``True`` in Python, so a JSON reply that (against the
    schema) returns a *string* would silently invert the flag. Accept real
    booleans and the canonical string spellings; fall back to ``default``.
    """
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        v = value.strip().lower()
        if v in ("true", "yes", "1"):
            return True
        if v in ("false", "no", "0"):
            return False
    return default


def neighbour_context(neighbour_text: str, *, is_previous: bool, at_book_edge: bool) -> str:
    """Describe a neighbouring story page for the clean/judge context (#73 M4).

    The boundary is derived from POSITION, not text truthiness: empty text at a
    true book edge is a genuine first/last page, but empty text at an INTERIOR
    neighbour is a wordless spread or a failed extraction — distinguishing them
    stops a run of missing pages each looking like the story's start/end (which
    let them all wrongly pass ``fits_context``).
    """
    if neighbour_text:
        return neighbour_text
    if at_book_edge:
        return (
            "(none — this is the first story page of the book)"
            if is_previous
            else "(none — this is the last story page of the book)"
        )
    return (
        "(the previous story page has no text — either a wordless spread or a "
        "failed extraction)"
        if is_previous
        else "(the next story page has no text — either a wordless spread or a "
        "failed extraction)"
    )


def sha1_hex(data: object) -> str:
    """SHA-1 hex digest of bytes or text (result-cache keys, #73 S9)."""
    if isinstance(data, bytes):
        return hashlib.sha1(data).hexdigest()
    return hashlib.sha1(str(data).encode("utf-8")).hexdigest()


def map_character_gender(value: object) -> str:
    """Map the DB character-gender code to a CharacterForm option.

    ``F`` -> "Female", ``M`` -> "Male", ``NGS`` -> "Non-specific". Any other
    value (the app has no "Non-binary" character option) falls back to
    "Non-specific".
    """
    code = str(value or "").strip().upper()
    if code == "F":
        return "Female"
    if code == "M":
        return "Male"
    # NGS ("no gender specified") and anything unexpected -> the indeterminate
    # option the app offers.
    return "Non-specific"


def map_human(value: object) -> tuple[bool, bool]:
    """Map the DB ``human`` code to (is_human, recognised).

    Whitespace is stripped first. ``H`` -> (True, True); ``NH`` -> (False, True).
    Any other value (the DB holds two strays, "NO" and "NGS") returns
    (True, False) — the Character.human default — with ``recognised=False`` so
    the caller can flag it.
    """
    code = str(value or "").strip().upper()
    if code == "H":
        return True, True
    if code == "NH":
        return False, True
    return True, False


def parse_year(value: object) -> Optional[int]:
    """Parse a 4-digit publication year in the app's accepted range (1900..).

    Returns None when absent or out of range. The Book form only offers years
    1900..current, so out-of-range values are treated as missing.
    """
    if value is None:
        return None
    text = str(value).strip()
    match = re.search(r"(\d{4})", text)
    if not match:
        return None
    year = int(match.group(1))
    if 1900 <= year <= datetime.now().year:
        return year
    return None


def parse_page_range(start: object, end: object) -> Optional[tuple[int, int]]:
    """Return (start, end) 1-based story page range, or None if unusable."""
    def _as_int(v: object) -> Optional[int]:
        if v is None:
            return None
        try:
            return int(float(str(v).strip()))
        except (TypeError, ValueError):
            return None

    s, e = _as_int(start), _as_int(end)
    if s is None or e is None:
        return None
    if s < 1 or e < s:
        return None
    return s, e


def is_story_page(page_number: int, page_range: Optional[tuple[int, int]]) -> bool:
    """True if the page counts as a story page.

    When a story range is present, pages inside ``[start, end]`` are story
    pages; when no range is known, every page is treated as a story page.
    """
    if page_range is None:
        return True
    start, end = page_range
    return start <= page_number <= end


def flanked_by_text_layer(pos: int, n_story: int, story_is_text_layer) -> bool:
    """True if the story page at story-position ``pos`` is safe to consider for
    a neighbour-continuity OCR skip.

    This is the pure classification the importer uses before ever calling the
    continuity judge (so it is unit-testable without any AI). ``story_is_text_layer``
    is a sequence, indexed by story-position, of booleans: True where that story
    page has a usable PDF text layer (>= ``TEXT_LAYER_MIN_CHARS``). The page at
    ``pos`` qualifies ONLY when it is an INTERIOR story page whose immediate story
    neighbours on BOTH sides are text-layer pages. It returns False (→ the caller
    OCRs unconditionally) for:

    * a story-range EDGE (first or last story page — no neighbour on one side);
    * a page whose previous OR next story neighbour is itself image-only, which
      also covers a RUN of consecutive image-only pages (each has an image-only
      neighbour), so a wordless run is never skipped on inferred context.

    The caller separately requires the neighbours to be TRUE text-layer text (not
    OCR'd) — which this guarantees, since text-layer pages are read from the PDF
    layer, never OCR — matching exactly what was validated.
    """
    if pos <= 0 or pos >= n_story - 1:
        return False
    return bool(story_is_text_layer[pos - 1]) and bool(story_is_text_layer[pos + 1])


def _compare_tokens(text: object) -> Counter:
    """Lower-case word multiset used for order-independent text comparison.

    Strips everything but alphanumerics so extraction/OCR whitespace, line-break
    and punctuation differences don't distort the overlap score.
    """
    words = re.sub(r"[^a-z0-9]+", " ", str(text or "").lower()).split()
    return Counter(words)


def text_similarity(a: object, b: object) -> float:
    """Word-overlap (Jaccard-on-multisets) similarity of two texts in [0, 1].

    Order-independent and cheap (no O(n^2) diff), so it scales to full-book
    text across the whole corpus. 1.0 == identical word bags; 0.0 == disjoint.
    Two empty texts count as identical; one empty as disjoint.
    """
    aw, bw = _compare_tokens(a), _compare_tokens(b)
    if not aw and not bw:
        return 1.0
    if not aw or not bw:
        return 0.0
    intersection = sum((aw & bw).values())
    union = sum((aw | bw).values())
    return intersection / union if union else 0.0


def text_containment(extracted: object, reference: object) -> float:
    """Asymmetric CONTAINMENT / recall of ``reference`` inside ``extracted`` (#73 S3).

    Returns ``sum(extracted & reference multiset) / sum(reference multiset)`` in
    [0, 1] — "what fraction of the validated reference's words did we capture".
    Unlike :func:`text_similarity` (symmetric Jaccard) this does NOT penalise
    extraction that legitimately reads MORE than the sparse validated pickle, so
    it is the primary flagging metric. An empty reference counts as fully
    contained (1.0 — there is nothing to capture); a non-empty reference against
    empty extracted text is 0.0.
    """
    ew, rw = _compare_tokens(extracted), _compare_tokens(reference)
    ref_total = sum(rw.values())
    if not ref_total:
        return 1.0
    captured = sum((ew & rw).values())
    return captured / ref_total


def normalise_blank_text(text: object) -> str:
    """Collapse quote/punctuation-only 'text' to a true empty string.

    OCR/extraction sometimes yields a page whose only content is stray
    punctuation or literal quote characters (e.g. a model that types ``""`` for
    a blank page). Such a page has no real words, so it is normalised to ``""``
    — which keeps the stored ``text`` clean and lets blank-page detection fire
    consistently. Any text containing at least one alphanumeric character is
    returned unchanged.
    """
    if not text:
        return ""
    text = str(text)
    # Ignore HTML-ish tags (e.g. a model emitting "<br>" for a line break) when
    # deciding whether any real content remains.
    without_tags = re.sub(r"<[^>]*>", "", text)
    return text if any(ch.isalnum() for ch in without_tags) else ""


def _alnum_tokens(text: object) -> list:
    """Alphanumeric-token list (multiset, case-sensitive) for the clean guard."""
    return re.findall(r"[0-9A-Za-z]+", str(text or ""))


def clean_kept(original: str, cleaned: str) -> bool:
    """True if a cleaned page only REMOVED content and may be kept (#73 S4).

    The clean pass is allowed to strip arbitrary junk but must never add or alter
    a word ("only remove, never change"). The correct invariant is therefore a
    token-subset check: every alphanumeric token (multiset) of the CLEANED text
    must already appear in the ORIGINAL. This accepts arbitrary removal and
    rejects any added/altered/case-changed word — much tighter than a
    char-ratio. An empty/blank clean is rejected (keep the original).
    """
    if not cleaned.strip():
        return False
    added = Counter(_alnum_tokens(cleaned)) - Counter(_alnum_tokens(original))
    return not added


# ---------------------------------------------------------------------------
# Firestore-reference placeholder (keeps document builders pure / testable).
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class Ref:
    """A pending Firestore reference to ``collection/doc_id``.

    Document builders emit these instead of real ``DocumentReference`` objects
    so they stay pure and testable without google-cloud installed. The live
    backend resolves each ``Ref`` to a real reference at write time; the
    dry-run backend renders it as ``collection/doc_id``.
    """

    collection: str
    doc_id: str

    def __str__(self) -> str:  # pragma: no cover - display only
        return f"{self.collection}/{self.doc_id}"


# ---------------------------------------------------------------------------
# Document builders — replicate each data_structures/*.py Firestore schema.
# ---------------------------------------------------------------------------

def build_author_doc(forename: str, surname: str, gender: str, now: datetime) -> dict:
    """Build an ``authors`` document (mirrors data_structures/author.py)."""
    return {
        "is_registered": True,
        "forename": forename,
        "surname": surname,
        "gender": gender,
        "entered_by": ENTERED_BY,
        "datetime_created": now,
        "last_updated": now,
    }


def build_named_doc(name: str, now: datetime) -> dict:
    """Build an ``illustrators`` / ``publishers`` single-name document."""
    return {
        "is_registered": True,
        "name": name,
        "entered_by": ENTERED_BY,
        "datetime_created": now,
        "last_updated": now,
    }


def build_book_doc(
    *,
    title: str,
    author_ref: Optional[Ref],
    illustrator_ref: Optional[Ref],
    publisher_ref: Optional[Ref],
    published: Optional[int],
    page_range: Optional[tuple[int, int]],
    page_count: int,
    character_refs: list,
    now: datetime,
    photos_uploaded: bool = True,
    needs_review: bool = False,
    review_pages: Optional[list] = None,
    high_priority_review: bool = False,
    review_note: str = "",
) -> dict:
    """Build a ``books`` document (mirrors data_structures/book.py Book.fields).

    Theme flags are intentionally omitted (all default to False on read via the
    app's backward-compatibility fallback) because the pilot data has no theme
    information — see PILOT_IMPORT.md.

    ``photos_uploaded`` reflects whether a PDF/pages actually exist (#73 M3):
    5 corpus books have no PDF, so claiming photos for them would be false —
    ``photos_url`` is left blank and the book is flagged. ``needs_review`` /
    ``review_pages`` / ``high_priority_review`` / ``review_note`` are additive
    diagnostic fields: ``needs_review`` is True (with the list of flagged page
    numbers and a short note) when the clean+judge pass flagged any page, or when
    the book has no PDF / no characters / an ambiguous author. The app's ``Book``
    model ignores undeclared fields, so this is safe and lets flagged books be
    found with a single Firestore query.
    """
    start, end = (page_range or (-1, -1))
    return {
        "is_registered": True,
        "title": title,
        "author": author_ref,
        "character_count": len(character_refs),
        "page_count": page_count,
        "word_count": -1,
        "sentence_count": -1,
        "datetime_created": now,
        "entered_by": ENTERED_BY,
        "entry_status": "complete",
        "first_content_page": start,
        "last_content_page": end,
        "illustrator": illustrator_ref,
        "publisher": publisher_ref,
        "last_updated": now,
        "published": published if published is not None else -1,
        "validated": True,
        "validated_by": ENTERED_BY,
        "photos_uploaded": bool(photos_uploaded),
        "photos_url": book_photos_url(title) if photos_uploaded else "",
        "comment": "",
        "datetime_submitted": now,
        "characters": list(character_refs),
        "needs_review": needs_review,
        "review_pages": list(review_pages or []),
        "high_priority_review": high_priority_review,
        "review_note": review_note,
    }


def build_page_doc(
    *,
    book_ref: Ref,
    page_number: int,
    contains_story: bool,
    text: str,
    now: datetime,
    needs_review: bool = False,
    review_note: str = "",
    review_priority: str = "",
    text_source: str = "none",
    clean_status: str = "",
    ocr_model: str = "",
    clean_model: str = "",
    continuity: Optional[dict] = None,
) -> dict:
    """Build a ``pages`` document (mirrors data_structures/page.py).

    ``needs_review`` / ``review_priority`` / ``review_note`` are extra diagnostic
    fields set by the clean+judge pass when the page text doesn't read as coherent
    story text or doesn't fit its neighbours — a candidate for re-extraction or
    human reading. ``review_priority`` is ``"high"`` when the page fails both the
    coherence and context-fit checks.

    ``text_source`` (#73 S8: ``"layer"`` | ``"ocr"`` | ``"skipped_wordless"`` |
    ``"none"``), ``clean_status`` (``"cleaned"`` / ``"unchanged"`` / ``"rejected"``
    / ``"failed"`` / ``""``) and the ``ocr_model`` / ``clean_model`` ids record how
    this page's text was produced and cleaned, so the OCR/AI subset is auditable
    and re-runnable. ``continuity`` (when set) records the neighbour-continuity
    judge verdict that decided an image-only page was a genuine wordless spread
    and its OCR was skipped (``text_source="skipped_wordless"``) — so the skip
    decision is auditable and re-runnable. All of these are additive: the app's
    ``Page`` model ignores fields it doesn't declare, so recording them here is
    safe and makes the flagged pages queryable in Firestore.
    """
    return {
        "is_registered": True,
        "book": book_ref,
        "page_number": page_number,
        "contains_story": contains_story,
        "text": text,
        "datetime_created": now,
        "entered_by": ENTERED_BY,
        "last_updated": now,
        "needs_review": needs_review,
        "review_priority": review_priority,
        "review_note": review_note,
        "text_source": text_source,
        "clean_status": clean_status,
        "ocr_model": ocr_model,
        "clean_model": clean_model,
        "continuity": dict(continuity) if continuity else None,
    }


def build_character_doc(
    *,
    book_ref: Ref,
    name: str,
    gender: str,
    protagonist: bool,
    human: bool,
    now: datetime,
) -> dict:
    """Build a ``characters`` document (mirrors data_structures/character.py)."""
    return {
        "is_registered": True,
        "book": book_ref,
        "name": name,
        "gender": gender,
        "ethnicity": DEFAULT_ETHNICITY,
        "disability": DEFAULT_DISABILITY,
        "protagonist": protagonist,
        "human": human,
        "plural": False,  # no plural/singular indicator in the pilot sources
        "datetime_created": now,
        "last_updated": now,
        "entered_by": ENTERED_BY,
    }


def build_alias_doc(
    *, character_ref: Ref, book_ref: Ref, name: str, now: datetime
) -> dict:
    """Build an ``aliases`` document (mirrors data_structures/alias.py)."""
    return {
        "is_registered": True,
        "character": character_ref,
        "book": book_ref,
        "name": name,
        "datetime_created": now,
        "last_updated": now,
        "entered_by": ENTERED_BY,
    }


# ---------------------------------------------------------------------------
# Source loading.
# ---------------------------------------------------------------------------

@dataclass
class ExcelBook:
    title: str
    norm: str
    author: str
    author_gender_code: str
    published: Optional[int]
    page_range: Optional[tuple[int, int]]
    row_number: int


def load_excel_books(path: str) -> list:
    """Load Sheet1 of the book-list workbook (openpyxl, data_only)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}

    title_i = idx.get("Title")
    # The author column has a trailing space ("Author ") in the source.
    author_i = idx.get("Author ", idx.get("Author"))
    gender_i = idx.get("Author Gender")
    start_i = idx.get("Starting Page")
    end_i = idx.get("Ending Page")
    year_i = idx.get("Year of First Publication")

    books: list = []
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        title = row[title_i] if title_i is not None else None
        if not title or not str(title).strip():
            continue
        title = str(title).strip()
        books.append(
            ExcelBook(
                title=title,
                norm=normalise_title(title),
                author=str(row[author_i]).strip() if author_i is not None and row[author_i] else "",
                author_gender_code=str(row[gender_i]).strip() if gender_i is not None and row[gender_i] else "",
                published=parse_year(row[year_i]) if year_i is not None else None,
                page_range=parse_page_range(
                    row[start_i] if start_i is not None else None,
                    row[end_i] if end_i is not None else None,
                ),
                row_number=row_number,
            )
        )
    wb.close()
    return books


@dataclass
class DbData:
    characters_by_book: dict  # norm title -> [character rows]
    aliases_by_book: dict     # norm title -> [alias rows]
    index_to_character: dict  # char index -> (norm book, name)
    human_anomalies: list     # rows with unrecognised human code


def load_db(path: str) -> DbData:
    """Load characters + aliases from the SQLite character database."""
    con = sqlite3.connect(path)
    con.row_factory = sqlite3.Row
    characters_by_book: dict = defaultdict(list)
    index_to_character: dict = {}
    human_anomalies: list = []
    for r in con.execute("SELECT * FROM characters"):
        book_norm = normalise_title(r["book"])
        name = str(r["name"]).strip()
        row = {
            "index": r["index"],
            "book_norm": book_norm,
            "book_raw": r["book"],
            "name": name,
            "gender": r["gender"],
            "human": r["human"],
            "is_protagonist": r["is_protagonist"],
        }
        characters_by_book[book_norm].append(row)
        index_to_character[r["index"]] = (book_norm, name)
        _human, recognised = map_human(r["human"])
        if not recognised:
            human_anomalies.append(row)

    aliases_by_book: dict = defaultdict(list)
    for r in con.execute("SELECT * FROM aliases"):
        book_norm = normalise_title(r["book"])
        aliases_by_book[book_norm].append(
            {
                "index": r["index"],
                "alias": str(r["alias"]).strip(),
                "character": str(r["character"]).strip(),
                "character_id": r["character_id"],
                "book_norm": book_norm,
                "book_raw": r["book"],
            }
        )
    con.close()
    return DbData(
        characters_by_book=dict(characters_by_book),
        aliases_by_book=dict(aliases_by_book),
        index_to_character=index_to_character,
        human_anomalies=human_anomalies,
    )


def load_pdfs(pdf_dir: str) -> tuple:
    """Return ({norm title: pdf path}, [duplicate groups]).

    ``.doc.pdf`` duplicates (e.g. ``Sing A Song Of Bottoms.doc.pdf``) collapse
    to the same normalised key; the plain ``.pdf`` wins and the duplicate is
    reported.
    """
    mapping: dict = {}
    groups: dict = defaultdict(list)
    if not os.path.isdir(pdf_dir):
        return mapping, []
    for entry in sorted(os.listdir(pdf_dir)):
        if not entry.lower().endswith(".pdf"):
            continue
        base = entry[:-4]
        base = re.sub(r"\.doc$", "", base)  # collapse "X.doc.pdf" -> "X"
        norm = normalise_title(base)
        groups[norm].append(entry)
        full = os.path.join(pdf_dir, entry)
        # Prefer the non-".doc" filename when both exist.
        if norm not in mapping or not entry.lower().endswith(".doc.pdf"):
            mapping[norm] = full
    dups = [(norm, files) for norm, files in groups.items() if len(files) > 1]
    return mapping, dups


def load_json_text(path: str) -> dict:
    """Return {norm title: full book text} from book_dataframe.json (fallback)."""
    if not os.path.exists(path):
        return {}
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    titles = data.get("Title", {})
    texts = data.get("Text", {})
    out: dict = {}
    for key, title in titles.items():
        out[normalise_title(title)] = texts.get(key, "")
    return out


def load_pickle_text(path: str) -> dict:
    """Return {norm title: full book text} from book_dataframe.pickle.

    This is the human-validated analysis text (issue #73) — the reference the
    per-page extraction is cross-checked against. Requires pandas; raises on any
    read/unpickle failure so the caller can fall back to the JSON export.
    """
    import pandas as pd  # lazy: only needed when a pickle reference is used

    df = pd.read_pickle(path)
    out: dict = {}
    for _, row in df.iterrows():
        out[normalise_title(row["Title"])] = str(row["Text"] or "")
    return out


def load_validated_text(pickle_path: str, json_path: str) -> tuple[dict, str]:
    """Load the per-book validated text, preferring the pickle over the JSON.

    Returns ``(mapping, source)`` where ``source`` is ``"pickle"`` or ``"json"``
    (or ``"none"`` if neither is available). The pickle is the human-validated
    analysis text; the JSON is a byte-identical export kept only as a fallback in
    case the pickle can't be unpickled on this interpreter.
    """
    if os.path.exists(pickle_path):
        try:
            mapping = load_pickle_text(pickle_path)
            if mapping:
                return mapping, "pickle"
        except Exception as exc:  # noqa: BLE001 - pandas/pickle version issues; fall back
            print(
                f"WARNING: could not read validated pickle {pickle_path} "
                f"({type(exc).__name__}: {exc}); falling back to JSON export.",
                file=sys.stderr,
            )
    json_map = load_json_text(json_path)
    return json_map, ("json" if json_map else "none")


# ---------------------------------------------------------------------------
# PDF processing (rendering via pypdfium2, text layer via pypdf).
# ---------------------------------------------------------------------------

def analyse_pdf(path: str) -> dict:
    """Return {page_count, text_layer_pages, image_only_pages, page_texts}.

    Uses pypdf's text-layer extraction only (no rendering) so it is cheap
    enough to run in the dry-run report. ``page_texts`` is a list of the
    extracted per-page text (may be blank for image-only pages).
    """
    from pypdf import PdfReader
    from pypdf.errors import PdfReadError

    reader = PdfReader(path)
    page_texts: list = []
    text_layer = 0
    for page in reader.pages:
        try:
            text = (page.extract_text() or "").strip()
        except (PdfReadError, KeyError, ValueError, TypeError):
            # pypdf raises assorted parse errors on malformed content streams;
            # treat an unreadable page's text layer as empty (it will OCR).
            text = ""
        page_texts.append(text)
        if len(text) >= TEXT_LAYER_MIN_CHARS:
            text_layer += 1
    return {
        "page_count": len(page_texts),
        "text_layer_pages": text_layer,
        "image_only_pages": len(page_texts) - text_layer,
        "page_texts": page_texts,
    }


def render_page_jpeg(pdf_doc, page_index: int, max_edge: int, quality: int) -> bytes:
    """Render one PDF page to JPEG bytes at ``max_edge`` longest edge (pypdfium2)."""
    page = pdf_doc[page_index]
    width, height = page.get_size()  # points
    scale = max_edge / max(width, height) if max(width, height) else 1.0
    bitmap = page.render(scale=scale)
    pil = bitmap.to_pil().convert("RGB")
    buf = io.BytesIO()
    pil.save(buf, format="JPEG", quality=quality)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# AI helpers (Anthropic) — best effort, degrade gracefully.
# ---------------------------------------------------------------------------

def _extract_json_object(text: str) -> Optional[dict]:
    """Pull the first JSON object out of a model reply, or None (fallback path).

    Only used when structured outputs are unavailable (older SDK/model) or the
    web-search lookup call, which cannot use structured outputs. Uses
    ``strict=False`` so a literal newline inside a JSON string value (which
    ``json.loads`` rejects by default as a control character) does not defeat the
    parse (#73 M6).
    """
    if not text:
        return None
    fenced = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.DOTALL)
    candidate = fenced.group(1) if fenced else None
    if candidate is None:
        brace = re.search(r"\{.*\}", text, re.DOTALL)
        candidate = brace.group(0) if brace else None
    if candidate is None:
        return None
    try:
        parsed = json.loads(candidate, strict=False)
        return parsed if isinstance(parsed, dict) else None
    except json.JSONDecodeError:
        return None


#: Shared rule appended to the OCR and clean/judge prompts (#73 M6): keep JSON
#: string values single-line so a stray literal newline can't break parsing.
_JSON_NEWLINE_RULE = (
    "Inside the JSON string values, encode any line break as \\n; never emit a "
    "literal newline inside a quoted string."
)

#: JSON schema for the OCR reply (structured outputs, #73 M6).
OCR_SCHEMA = {
    "type": "object",
    "properties": {
        "has_text": {"type": "boolean"},
        "text": {"type": "string"},
    },
    "required": ["has_text", "text"],
    "additionalProperties": False,
}

#: JSON schema for the clean/judge reply (structured outputs, #73 M6).
CLEAN_SCHEMA = {
    "type": "object",
    "properties": {
        "cleaned_text": {"type": "string"},
        "makes_sense": {"type": "boolean"},
        "fits_context": {"type": "boolean"},
        "review_note": {"type": "string"},
    },
    "required": ["cleaned_text", "makes_sense", "fits_context", "review_note"],
    "additionalProperties": False,
}

#: Static (cacheable) OCR instruction prefix. Placed FIRST in the content list
#: with cache_control so the identical prefix can be prompt-cached across every
#: page; the variable image block follows it (#73 prompt caching).
OCR_STATIC_PROMPT = (
    "You are transcribing the text from a single page of a printed "
    "children's picture book.\n\n"
    "Transcribe ALL of the book's OWN text that appears on this page, exactly "
    "as printed, with this guidance:\n"
    "- Picture-book typography is often irregular: decorative, hand-drawn or "
    "unusual fonts, widely varying sizes, coloured text over busy "
    "illustrations, text that curves or sits at an angle, speech bubbles, and "
    "large sound-effect words (e.g. 'SPLASH!', 'WHOOSH', 'Ker-BOOM'). Read "
    "all of it.\n"
    "- The page or its text may be rotated or skewed; read it in its correct "
    "upright reading orientation.\n"
    "- Transcribe the exact wording, spelling, capitalisation and punctuation "
    "as printed. Children's books deliberately use invented words, made-up "
    "names, onomatopoeia and non-standard spelling — preserve these verbatim; "
    "do NOT correct or standardise them.\n"
    "- Preserve the reading order and keep line breaks roughly as printed.\n"
    "- Transcribe ONLY the book's narrative text (story, dialogue, captions, "
    "sound words). Do NOT transcribe page numbers or running heads, publisher "
    "/ copyright / ISBN / barcode text, or text that is part of the "
    "illustration artwork (e.g. signposts, shop names, labels).\n\n"
    "Decide whether the page contains any of the book's own text at all. Many "
    "picture-book pages are WORDLESS full illustrations — that is normal and "
    "expected.\n\n"
    "Respond with ONLY a JSON object, no code fences and no commentary:\n"
    '{"has_text": true or false, "text": "<the exact transcription, or an '
    'empty string>"}\n\n'
    "If the page has no book text (a wordless illustration), set "
    '"has_text" to false and "text" to "". NEVER put a description, '
    "explanation or note about the page into the text field — the text field "
    "must contain only the book's transcribed words and nothing else.\n\n"
    + _JSON_NEWLINE_RULE
)

#: Static (cacheable) clean/judge instruction prefix. Placed FIRST with
#: cache_control; the variable context + this-page block follows it (#73 prompt
#: caching, M4/S5/S6).
CLEAN_STATIC_PROMPT = (
    "You are given the extracted text (via a PDF text layer or OCR) of a single "
    "page of a children's picture book, together with the previous and next "
    "pages' text as context only. Do THREE things.\n\n"
    "1. CLEAN this page's text. Remove ONLY content that is not the story's own "
    "words: stray garbage character sequences (for example 'as&ij-'), OCR "
    "noise, control characters, mojibake, and broken or duplicated fragments; "
    "AND print artefacts such as bare page numbers, running heads (the book "
    "title or chapter repeated at the top or bottom of the page), and "
    "publisher / copyright / ISBN boilerplate. Do NOT correct spelling, "
    "grammar, punctuation or capitalisation, and do NOT change unusual, made-up "
    "or oddly spelled words — children's books deliberately use non-standard "
    "spelling, invented words and playful sounds, and every real story word "
    "MUST be preserved exactly as written. Preserve line breaks and the reading "
    "order. If it is already clean, keep it unchanged.\n\n"
    "2. Judge makes_sense (a property of THIS PAGE on its own). This is TRUE "
    "when the cleaned text reads as genuine, intact language from a "
    "children's picture-book story: it forms real, readable words and "
    "phrases that a person could read aloud as part of a story — even if it "
    "is very short, sparse, rhyming, repetitive, playful, a single line of "
    "dialogue, or just a big sound-effect word like 'SPLASH!'. A page that ends "
    "mid-sentence because the sentence continues on the next page is completely "
    "normal and DOES make sense. It is FALSE only when the text is not usable "
    "story text: garbled or scrambled characters, random disconnected "
    "word-fragments, obvious OCR nonsense or mojibake, text duplicated or "
    "looping on itself, or an editorial description/note ABOUT the page (e.g. "
    "'this page is a wordless illustration') rather than the story's own words. "
    "An EMPTY page (a wordless illustration — very common and completely normal "
    "in picture books) is not garbled, so treat empty text as "
    "makes_sense=true.\n\n"
    "3. Judge fits_context (how THIS PAGE sits between its neighbours). This "
    "is TRUE when the page belongs naturally in the story at this point: the "
    "narrative reads continuously from the previous page, through this page, "
    "into the next page, allowing for the normal way picture books work — "
    "page turns, scene changes, a new character or setting being introduced, "
    "and refrains or repeated phrases that recur across pages. It is FALSE "
    "when this page clearly does not belong here: the text reads as if it "
    "came from a different book or a different part of the story, it "
    "introduces characters or events that contradict the surrounding pages, "
    "it abruptly repeats or duplicates a neighbouring page's wording, or the "
    "page is EMPTY at a point where the story plainly continues so that text "
    "appears to be MISSING (an OCR miss). If this page is empty but the "
    "surrounding pages still flow together sensibly across it (a genuine "
    "wordless spread), that DOES fit.\n\n"
    "Respond with ONLY a JSON object (no code fences, no commentary):\n"
    '{"cleaned_text": "<this page cleaned>", "makes_sense": true or false, '
    '"fits_context": true or false, "review_note": "<short reason if either '
    'is false, else empty>"}\n\n'
    + _JSON_NEWLINE_RULE
)


def _supports_structured_outputs(client) -> bool:
    """True if this SDK's ``messages.create`` accepts ``output_config`` (#73 M6)."""
    try:
        sig = inspect.signature(client.messages.create)
    except (TypeError, ValueError):
        return False
    return "output_config" in sig.parameters


def _strip_cache_control(blocks: list) -> list:
    """Return content blocks with any ``cache_control`` key removed (SDK fallback)."""
    out: list = []
    for b in blocks:
        if isinstance(b, dict) and "cache_control" in b:
            b = {k: v for k, v in b.items() if k != "cache_control"}
        out.append(b)
    return out


def _ai_json_call(
    client, *, model: str, max_tokens: int, content_blocks: list, schema: dict
) -> tuple[Optional[dict], str]:
    """Make a Claude call expecting a JSON object; return ``(parsed_or_None, raw)``.

    Prefers Anthropic structured outputs (``output_config`` json_schema) so the
    reply is guaranteed valid JSON — no fence/regex parsing needed for the OCR
    and clean/judge calls. If the SDK is too old to accept ``output_config`` /
    ``cache_control`` it retries plain and falls back to manual JSON extraction.

    Prompt caching (``cache_control`` on the leading static block of
    ``content_blocks``) only takes effect once the cached prefix exceeds the
    model's minimum (~1024 tokens); it is harmless below that. The Anthropic API
    error PROPAGATES so the caller can record a high-priority flag (#73 M1); a
    ``None`` dict means valid JSON could not be parsed (only on the manual
    fallback path).
    """
    messages = [{"role": "user", "content": content_blocks}]
    use_structured = _supports_structured_outputs(client)
    try:
        if use_structured:
            response = client.messages.create(
                model=model,
                max_tokens=max_tokens,
                messages=messages,
                output_config={"format": {"type": "json_schema", "schema": schema}},
            )
        else:
            response = client.messages.create(
                model=model, max_tokens=max_tokens, messages=messages
            )
    except TypeError:
        # Older SDK: output_config and/or cache_control not accepted — retry plain.
        use_structured = False
        response = client.messages.create(
            model=model,
            max_tokens=max_tokens,
            messages=[{"role": "user", "content": _strip_cache_control(content_blocks)}],
        )
    raw = "".join(
        block.text for block in response.content if getattr(block, "type", None) == "text"
    ).strip()
    if use_structured:
        try:
            parsed = json.loads(raw, strict=False)
        except json.JSONDecodeError:
            return _extract_json_object(raw), raw
        return (parsed if isinstance(parsed, dict) else None), raw
    return _extract_json_object(raw), raw


def ai_lookup_book_metadata(client, title: str, author: str, model: str) -> dict:
    """Look up illustrator / publisher / year for a book via Claude + web search.

    Returns ``{"illustrator": str, "publisher": str, "year": Optional[int]}``,
    using "Unknown" for anything not confidently found. UK-first (#73 S1): this
    is a UK primary-school corpus, so the prompt asks for the ORIGINAL UK
    publisher/imprint as printed (e.g. Walker Books, not its US partner
    Candlewick Press), the illustrator as credited, and the year of FIRST
    publication. Any failure degrades to all-"Unknown" rather than raising — this
    is best-effort enrichment. (Web search precludes structured outputs, so this
    call keeps manual JSON parsing.)
    """
    static_instructions = (
        "You are cataloguing a children's picture book from a UK primary-school "
        "reading corpus. Using web search, identify, for this SPECIFIC book:\n"
        "- illustrator: the illustrator as credited on the book (if the author "
        "illustrated it themselves, use the author's name);\n"
        "- publisher: the publisher of the FIRST edition, preferring the original "
        "UK publisher/imprint as printed on the book. This is a UK corpus, so "
        "prefer the UK imprint over a US partner (for example Walker Books, NOT "
        "its US partner Candlewick Press). Give the imprint, not the parent "
        "conglomerate;\n"
        "- year: the year of FIRST publication (not a later reissue or reprint).\n\n"
        'Reply with ONLY a JSON object of the form {"illustrator": "...", '
        '"publisher": "...", "year": 1999}. Use the string "Unknown" for any '
        "field you cannot confirm, and null for year if unknown. Do NOT guess.\n\n"
        + _JSON_NEWLINE_RULE
    )
    variable_block = f"Title: {title}\nAuthor: {author or 'unknown'}"
    # cache_control on the static instruction block so the identical prefix is
    # prompt-cached across every book (only bites above the model minimum).
    content = [
        {"type": "text", "text": static_instructions, "cache_control": {"type": "ephemeral"}},
        {"type": "text", "text": variable_block},
    ]
    tools = [{"type": "web_search_20260209", "name": "web_search"}]
    try:
        messages = [{"role": "user", "content": content}]
        response = client.messages.create(
            model=model, max_tokens=1024, tools=tools, messages=messages
        )
        continuations = 0
        while getattr(response, "stop_reason", None) == "pause_turn" and continuations < 3:
            messages.append({"role": "assistant", "content": response.content})
            response = client.messages.create(
                model=model, max_tokens=1024, tools=tools, messages=messages
            )
            continuations += 1
        text = "".join(
            block.text for block in response.content if getattr(block, "type", None) == "text"
        )
        data = _extract_json_object(text) or {}
    except Exception as exc:  # noqa: BLE001 - network/parse; degrade to Unknown, but surface why
        print(f"    [ai-lookup] failed for {title!r}: {type(exc).__name__}: {exc}", file=sys.stderr)
        data = {}

    illustrator = str(data.get("illustrator") or "Unknown").strip() or "Unknown"
    publisher = str(data.get("publisher") or "Unknown").strip() or "Unknown"
    year = parse_year(data.get("year"))
    return {"illustrator": illustrator, "publisher": publisher, "year": year}


def ai_ocr_page(client, jpeg_bytes: bytes, model: str) -> tuple[str, Optional[str]]:
    """OCR a rendered page image with Claude vision.

    Returns ``(text, error)``: ``error`` is ``None`` on success (``text`` may be
    ``""`` for a genuinely wordless page) and a short ``"OCR call failed: <Exc>"``
    string when the API call itself failed (#73 M1). This lets the caller tell a
    transient failure apart from a real blank page and flag the former
    ``needs_review`` high-priority instead of silently storing ``""``.

    Uses a structured reply — ``{"has_text": bool, "text": str}`` (#73 M6) — so a
    wordless page is declared explicitly rather than the model narrating "this
    page is a wordless illustration…". The static instruction block is sent FIRST
    (cache_control) followed by the variable image block for prompt caching.
    """
    import base64

    b64 = base64.standard_b64encode(jpeg_bytes).decode("utf-8")
    content = [
        {"type": "text", "text": OCR_STATIC_PROMPT, "cache_control": {"type": "ephemeral"}},
        {
            "type": "image",
            "source": {"type": "base64", "media_type": "image/jpeg", "data": b64},
        },
    ]
    try:
        data, raw = _ai_json_call(
            client, model=model, max_tokens=2048, content_blocks=content, schema=OCR_SCHEMA
        )
    except Exception as exc:  # noqa: BLE001 - network/parse; RECORD the failure (M1)
        print(f"    [ai-ocr] failed: {type(exc).__name__}: {exc}", file=sys.stderr)
        return "", f"OCR call failed: {type(exc).__name__}"

    if data is None:
        # Non-JSON reply (only reachable on the manual-parse fallback). Salvage
        # cautiously, and NEVER store a raw JSON-ish blob as page text (#73 M6).
        print("    [ai-ocr] non-JSON reply; salvaging cautiously", file=sys.stderr)
        stripped = raw.strip()
        lowered = raw.lower()
        if stripped.startswith("{") or "has_text" in lowered:
            return "", None
        if any(p in lowered for p in ("no text", "no book text", "wordless", "no visible text", "appears to be")):
            return "", None
        return normalise_blank_text(raw), None
    if not _as_bool(data.get("has_text"), False):
        return "", None
    return normalise_blank_text(str(data.get("text") or "")), None


def derive_review(makes_sense: bool, fits_context: bool, has_text: bool = True) -> tuple[bool, str]:
    """Turn the two judge booleans into ``(needs_review, priority)``.

    A page is flagged for review if it fails either check. Priority is **high**
    when it fails BOTH — i.e. the text neither reads as sensible story text on its
    own NOR fits the pages either side of it (the strongest signal that the
    extraction is wrong / text is missing). Additionally (#73 S7), an EMPTY page
    that does not fit its context is high-priority: an OCR-missed page reads
    makes_sense=true (empty isn't garbled) but breaks the narrative, and is the
    top re-extraction candidate. Failing just one check otherwise is ``normal``
    priority; passing both is not flagged.
    """
    if makes_sense and fits_context:
        return False, ""
    if not makes_sense and not fits_context:
        return True, "high"
    if not has_text and not fits_context:
        return True, "high"
    return True, "normal"


def ai_clean_and_judge(
    client, text: str, prev_context: str, next_context: str, model: str
) -> tuple[str, str, bool, str, str]:
    """Clean junk from page text AND judge it in context, in one call.

    One Claude call does three jobs for a single story page: (1) strip
    extraction/OCR noise and print artefacts (page numbers, running heads,
    copyright boilerplate) without altering any real story word; (2) judge
    ``makes_sense`` (coherent on its own); (3) judge ``fits_context`` given the
    neighbouring pages.

    ``prev_context`` / ``next_context`` are the already-resolved neighbour
    descriptors (see :func:`neighbour_context`, #73 M4): a real page's text, or a
    position-derived boundary/​missing-text note. Returns
    ``(result_text, status, needs_review, priority, review_note)`` where
    ``status`` is ``"cleaned"`` / ``"unchanged"`` / ``"rejected"`` / ``"failed"``.

    Failure handling (#73 M1/M5): an API error or unparseable reply is stored as
    ``needs_review=True`` (note "clean/judge call failed", normal priority) rather
    than silently unflagged; a clean rejected by the divergence guard forces
    ``needs_review=True`` (priority at least normal) because the ORIGINAL (possibly
    garbage) text is kept.
    """
    context_block = (
        "PREVIOUS PAGE TEXT (for context only, do not clean or return it):\n"
        f"{prev_context}\n\n"
        "NEXT PAGE TEXT (for context only, do not clean or return it):\n"
        f"{next_context}\n\n"
        "THIS PAGE TEXT:\n"
        + (text or "(this page has no extracted text)")
    )
    content = [
        {"type": "text", "text": CLEAN_STATIC_PROMPT, "cache_control": {"type": "ephemeral"}},
        {"type": "text", "text": context_block},
    ]
    try:
        data, _raw = _ai_json_call(
            client, model=model, max_tokens=2048, content_blocks=content, schema=CLEAN_SCHEMA
        )
    except Exception as exc:  # noqa: BLE001 - network/parse; RECORD the failure (M1)
        print(f"    [ai-clean] failed: {type(exc).__name__}: {exc}", file=sys.stderr)
        return text, "failed", True, "normal", "clean/judge call failed"

    if data is None:
        # No usable JSON came back; flag it rather than silently passing (M1).
        print("    [ai-clean] non-JSON reply; keeping original text", file=sys.stderr)
        return text, "failed", True, "normal", "clean/judge call failed"

    makes_sense = _as_bool(data.get("makes_sense", True), True)
    fits_context = _as_bool(data.get("fits_context", True), True)
    review_note = str(data.get("review_note") or "").strip()
    cleaned = str(data.get("cleaned_text") or "").strip()

    if not cleaned or cleaned == text:
        final_text, status = text, "unchanged"
    elif clean_kept(text, cleaned):
        final_text, status = cleaned, "cleaned"
    else:
        final_text, status = text, "rejected"

    needs_review, priority = derive_review(
        makes_sense, fits_context, has_text=bool(final_text.strip())
    )

    if status == "rejected":
        # The kept text is the ORIGINAL (possibly garbage); the verdict above was
        # for the discarded clean, so force a review flag (#73 M5).
        needs_review = True
        if priority != "high":
            priority = "normal"
        review_note = "AI clean diverged from original and was discarded; original kept"

    return final_text, status, needs_review, priority, review_note


# ---------------------------------------------------------------------------
# Local result cache (#73 S9) — cheap re-runs, never crashes the run.
# ---------------------------------------------------------------------------

class ResultCache:
    """A tiny per-key JSON file cache for OCR / clean+judge results (#73 S9).

    Keyed by content hash so a crash or an ``--overwrite`` re-run reuses prior
    API results instead of re-paying for them. Deliberately robust: a corrupt or
    missing cache file never raises — it simply misses. ``enabled=False`` (via
    ``--no-cache``) turns it into a no-op.
    """

    def __init__(self, cache_dir: str, enabled: bool = True):
        self.enabled = enabled and bool(cache_dir)
        self.cache_dir = cache_dir
        if self.enabled:
            try:
                os.makedirs(cache_dir, exist_ok=True)
            except OSError as exc:
                print(f"  (cache disabled: {type(exc).__name__}: {exc})", file=sys.stderr)
                self.enabled = False

    def _path(self, key: str) -> str:
        return os.path.join(self.cache_dir, sha1_hex(key) + ".json")

    def get(self, key: str) -> Optional[dict]:
        if not self.enabled:
            return None
        try:
            with open(self._path(key), encoding="utf-8") as fh:
                value = json.load(fh)
            return value if isinstance(value, dict) else None
        except (OSError, ValueError):
            return None

    def put(self, key: str, value: dict) -> None:
        if not self.enabled:
            return
        try:
            with open(self._path(key), "w", encoding="utf-8") as fh:
                json.dump(value, fh)
        except (OSError, TypeError, ValueError):
            pass  # caching is best-effort; never fail the run over it


class AICircuitBreakerError(RuntimeError):
    """Raised to abort the run after too many consecutive AI failures (#73 M1)."""


# ---------------------------------------------------------------------------
# Backends.
# ---------------------------------------------------------------------------

class Backend:
    """Interface for the two run modes (dry-run vs live)."""

    def document_exists(self, collection: str, doc_id: str) -> bool:
        raise NotImplementedError

    def resolve_ref(self, ref: Optional[Ref]):
        raise NotImplementedError

    def write_document(self, collection: str, doc_id: str, data: dict) -> None:
        raise NotImplementedError

    def s3_put(self, path: str, data: bytes) -> None:
        raise NotImplementedError


class DryRunBackend(Backend):
    """Writes nothing. ``document_exists`` reports absent so the plan is full."""

    def document_exists(self, collection: str, doc_id: str) -> bool:
        return False

    def resolve_ref(self, ref: Optional[Ref]):
        return ref  # rendered as "collection/id" by str()

    def write_document(self, collection: str, doc_id: str, data: dict) -> None:  # noqa: D401
        return None

    def s3_put(self, path: str, data: bytes) -> None:  # noqa: D401
        return None


class LiveBackend(Backend):
    """Real Firestore + S3 writes (mirrors data_cleanup.py's LiveBackend)."""

    def __init__(self, secrets: dict):
        import json as _json

        from google.cloud import firestore
        from google.oauth2 import service_account
        import s3fs

        firestore_key = secrets.get("firestore_key")
        if not firestore_key:
            raise KeyError("secrets is missing required 'firestore_key'")
        key_info = _json.loads(firestore_key) if isinstance(firestore_key, str) else firestore_key
        creds = service_account.Credentials.from_service_account_info(key_info)
        self._db = firestore.Client(credentials=creds, project=FIRESTORE_PROJECT)

        for required in ("AWS_ACCESS_KEY_ID", "AWS_SECRET_ACCESS_KEY"):
            if required not in secrets:
                raise KeyError(f"secrets is missing required '{required}'")
        self._fs = s3fs.S3FileSystem(
            anon=False,
            key=secrets["AWS_ACCESS_KEY_ID"],
            secret=secrets["AWS_SECRET_ACCESS_KEY"],
        )

    def document_exists(self, collection: str, doc_id: str) -> bool:
        return self._db.collection(collection).document(doc_id).get().exists

    def resolve_ref(self, ref: Optional[Ref]):
        if ref is None:
            return None
        return self._db.collection(ref.collection).document(ref.doc_id)

    def write_document(self, collection: str, doc_id: str, data: dict) -> None:
        # Convert any pending Refs (scalar or in a list) to real references.
        resolved = {}
        for key, value in data.items():
            if isinstance(value, Ref):
                resolved[key] = self.resolve_ref(value)
            elif isinstance(value, list):
                resolved[key] = [
                    self.resolve_ref(v) if isinstance(v, Ref) else v for v in value
                ]
            else:
                resolved[key] = value
        self._db.collection(collection).document(doc_id).set(resolved, merge=True)

    def s3_put(self, path: str, data: bytes) -> None:
        with self._fs.open(path, "wb") as fh:
            fh.write(data)


# ---------------------------------------------------------------------------
# Secrets loading (mirrors scripts/data_cleanup.py).
# ---------------------------------------------------------------------------

def load_secrets(path: str) -> dict:
    """Load ``.streamlit/secrets.toml`` directly (no Streamlit dependency)."""
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"secrets file not found: {path} (run from the project root, or pass "
            "--secrets). This tool must run where the real secrets exist."
        )
    try:
        import tomllib  # type: ignore

        with open(path, "rb") as f:
            return tomllib.load(f)
    except ModuleNotFoundError:
        pass
    try:
        import tomli  # type: ignore

        with open(path, "rb") as f:
            return tomli.load(f)
    except ModuleNotFoundError:
        pass
    import toml  # type: ignore

    with open(path, "r", encoding="utf-8") as f:
        return toml.load(f)


# ---------------------------------------------------------------------------
# Planning + execution.
# ---------------------------------------------------------------------------

@dataclass
class Totals:
    books: int = 0
    books_skipped: int = 0
    authors: set = field(default_factory=set)
    illustrators: set = field(default_factory=set)
    publishers: set = field(default_factory=set)
    pages: int = 0
    story_pages: int = 0
    images: int = 0
    characters: int = 0
    aliases: int = 0
    pages_text_layer: int = 0
    pages_needing_ocr: int = 0
    #: Image-only story pages whose OCR was SKIPPED because the neighbour-continuity
    #: judge found the story flows straight through (genuine wordless spread).
    pages_skipped_ocr: int = 0
    pages_cleaned: int = 0
    pages_clean_rejected: int = 0
    pages_flagged: int = 0
    pages_flagged_high: int = 0
    #: AI-call failure counters (#73 M1).
    ocr_failed: int = 0
    judge_failed: int = 0
    #: Shared-entity write collisions skipped to protect human data (#73 M2).
    authors_existing: int = 0
    illustrators_existing: int = 0
    publishers_existing: int = 0
    #: (title, [flagged page numbers], high_priority) for flagged books.
    books_flagged: list = field(default_factory=list)
    #: (title, containment, jaccard) for every book with a validated reference.
    similarities: list = field(default_factory=list)
    #: books with NO validated-text reference to cross-check against.
    no_reference: list = field(default_factory=list)


def write_shared_entity(
    backend: "Backend", collection: str, doc_id: str, doc: dict
) -> bool:
    """Write a shared entity doc ONLY if it does not already exist (#73 M2).

    ``authors`` / ``illustrators`` / ``publishers`` are shared with real
    human-entered data, so importing must never clobber a person's manually set
    gender/entered_by. Returns True if the write happened, False if the doc
    already existed (the reference is still used; the write is skipped and the
    collision counted by the caller). In dry-run ``document_exists`` always
    reports absent, so the plan shows every entity as written.
    """
    if backend.document_exists(collection, doc_id):
        return False
    backend.write_document(collection, doc_id, doc)
    return True


def plan_and_run(args) -> int:
    """Load sources, match, print the plan, and (with --execute) import."""
    methods = args.methods_dir
    excel_path = args.excel or os.path.join(methods, "Book-List-Final-NONA.xlsx")
    db_path = args.db or os.path.join(methods, "character_database.db")
    pdf_dir = args.pdf_dir or os.path.join(methods, "text_pdfs")
    json_path = args.json or os.path.join(methods, "data", "book_dataframe.json")
    pickle_path = args.pickle or os.path.join(methods, "data", "book_dataframe.pickle")

    for label, p in [("excel", excel_path), ("db", db_path)]:
        if not os.path.exists(p):
            print(f"ERROR: {label} not found: {p}", file=sys.stderr)
            return 2
    if not os.path.isdir(pdf_dir):
        # Not fatal: books without a matching PDF simply get no page images or
        # per-page text (this is also how individual missing-PDF books are
        # handled). Warn so the operator knows images won't be imported.
        print(f"WARNING: pdf dir not found ({pdf_dir}); no page images/pages "
              f"will be imported.", file=sys.stderr)

    print("=" * 78)
    print(f"PILOT IMPORT — {'EXECUTE (LIVE WRITES)' if args.execute else 'DRY RUN (no writes)'}")
    print("=" * 78)
    print(f"  excel : {excel_path}")
    print(f"  db    : {db_path}")
    print(f"  pdfs  : {pdf_dir}")
    print(f"  pickle: {pickle_path}")
    print(f"  json  : {json_path}")
    print()

    excel_books = load_excel_books(excel_path)
    db = load_db(db_path)
    pdf_map, pdf_dups = load_pdfs(pdf_dir)
    validated_text, validated_source = load_validated_text(pickle_path, json_path)
    print(f"  validated text source: {validated_source} ({len(validated_text)} books)")
    print()

    excel_norms = {b.norm for b in excel_books}
    db_norms = set(db.characters_by_book)
    pdf_norms = set(pdf_map)

    # --- Coverage / discrepancy report ------------------------------------
    print("SOURCE COVERAGE")
    print(f"  excel books (unique titles) : {len(excel_norms)}")
    print(f"  db books (with characters)  : {len(db_norms)}")
    print(f"  pdf files (unique titles)   : {len(pdf_norms)}")
    print(f"  validated text entries      : {len(validated_text)} (from {validated_source})")
    print()

    missing_pdf = sorted(b.title for b in excel_books if b.norm not in pdf_norms)
    missing_db = sorted(b.title for b in excel_books if b.norm not in db_norms)
    pdf_not_in_excel = sorted(pdf_map[n] for n in pdf_norms - excel_norms)
    db_not_in_excel = sorted(n for n in db_norms - excel_norms)
    pdf_not_in_db = sorted(pdf_map[n] for n in pdf_norms - db_norms)

    print("GAPS / DISCREPANCIES")
    print(f"  excel titles with NO pdf ({len(missing_pdf)}):")
    for t in missing_pdf:
        print(f"      - {t}")
    print(f"  excel titles with NO db characters ({len(missing_db)}):")
    for t in missing_db:
        print(f"      - {t}")
    print(f"  pdf files with NO excel row ({len(pdf_not_in_excel)}):")
    for t in pdf_not_in_excel:
        print(f"      - {os.path.basename(t)}")
    print(f"  db books with NO excel row ({len(db_not_in_excel)}):")
    for t in db_not_in_excel:
        print(f"      - {t}")
    print(f"  pdf files with NO db characters ({len(pdf_not_in_db)}):")
    for t in pdf_not_in_db:
        print(f"      - {os.path.basename(t)}")
    print(f"  duplicate pdf filename groups ({len(pdf_dups)}):")
    for norm, files in pdf_dups:
        print(f"      - {files}")
    print(f"  db 'human' anomalies (not H/NH) ({len(db.human_anomalies)}):")
    for row in db.human_anomalies:
        print(f"      - book={row['book_raw']!r} name={row['name']!r} human={row['human']!r}")
    multi_author = [
        b for b in excel_books
        if " and " in f" {b.author.lower()} " or "/" in b.author_gender_code
    ]
    print(f"  multi-author / ambiguous author-gender rows ({len(multi_author)}):")
    for b in multi_author:
        print(f"      - {b.title!r} author={b.author!r} gender_code={b.author_gender_code!r}")
    print(
        "  NOTE: character 'plural' has no source in the sheet or DB -> "
        "defaulted to False for every character."
    )
    print()

    # --- Backend + AI client ---------------------------------------------
    backend: Backend
    ai_client = None
    if args.execute:
        secrets = load_secrets(args.secrets)
        backend = LiveBackend(secrets)
        if "ANTHROPIC_API_KEY" in secrets:
            import anthropic

            ai_client = anthropic.Anthropic(api_key=secrets["ANTHROPIC_API_KEY"])
        else:
            print("WARNING: no ANTHROPIC_API_KEY in secrets; illustrator/publisher "
                  "lookup and OCR fallback will use 'Unknown'/blank.", file=sys.stderr)
    else:
        backend = DryRunBackend()
        # In dry-run we optionally build a client only to demonstrate a sample
        # lookup, using secrets if they happen to be present.
        if args.sample_ai and os.path.exists(args.secrets):
            try:
                secrets = load_secrets(args.secrets)
                if "ANTHROPIC_API_KEY" in secrets:
                    import anthropic

                    ai_client = anthropic.Anthropic(api_key=secrets["ANTHROPIC_API_KEY"])
            except Exception as exc:  # noqa: BLE001 - sample only; never fatal in dry-run
                print(f"  (sample AI unavailable: {type(exc).__name__}: {exc})")

    now = datetime.now(timezone.utc)
    totals = Totals()

    # Local result cache (#73 S9): reuse OCR / clean+judge results across runs.
    cache = ResultCache(args.cache_dir, enabled=not args.no_cache)
    if args.execute and cache.enabled:
        print(f"  result cache: {args.cache_dir}")
    # Circuit-breaker state (#73 M1): consecutive AI failures.
    consecutive_ai_failures = 0
    aborted = False

    matched = [
        b for b in excel_books
        if b.norm in pdf_norms or b.norm in db_norms or b.norm in excel_norms
    ]
    if args.limit:
        matched = matched[: args.limit]

    sample_ai_budget = args.sample_ai if not args.execute else 10 ** 9
    sample_ai_done = 0

    def _note_ai_failure() -> None:
        """Record one AI failure; abort the run if too many in a row (#73 M1)."""
        nonlocal consecutive_ai_failures
        consecutive_ai_failures += 1
        if consecutive_ai_failures >= MAX_CONSECUTIVE_AI_FAILURES:
            raise AICircuitBreakerError(
                f"{consecutive_ai_failures} consecutive AI failures "
                f"(>= {MAX_CONSECUTIVE_AI_FAILURES}); aborting. The run is "
                "resumable — book documents are written last, so this book is "
                "re-done and already-written books are skipped on the next run."
            )

    try:
      for book in matched:
        book_id = book_document_id(book.title)
        if not args.overwrite and backend.document_exists("books", book_id):
            totals.books_skipped += 1
            print(f"SKIP (exists): {book.title}  [{book_id}]")
            continue

        db_chars = db.characters_by_book.get(book.norm, [])
        db_aliases = db.aliases_by_book.get(book.norm, [])
        pdf_path = pdf_map.get(book.norm)

        # AI illustrator/publisher/year lookup (best effort). In dry-run only a
        # small sample is actually called; the rest are planned as 'Unknown'.
        do_lookup = ai_client is not None and (args.execute or sample_ai_done < sample_ai_budget)
        if do_lookup:
            meta = ai_lookup_book_metadata(ai_client, book.title, book.author, args.lookup_model)
            if not args.execute:
                sample_ai_done += 1
        else:
            meta = {"illustrator": "Unknown", "publisher": "Unknown", "year": None}

        published = book.published if book.published is not None else meta["year"]
        illustrator_name = meta["illustrator"]
        publisher_name = meta["publisher"]

        book_ref = Ref("books", book_id)

        # Book-level review notes aggregated across author/PDF/character checks.
        book_review_notes: list = []

        # Author — recover gender from "X and Y" multi-author rows (#73 S2).
        author_parse = parse_author_field(book.author, book.author_gender_code)
        author_gender = map_author_gender(author_parse.gender_code)
        if author_parse.needs_review and author_parse.note:
            book_review_notes.append(author_parse.note)
        author_ref = None
        if author_parse.author_name:
            forename, surname = split_name(author_parse.author_name)
            author_id = person_document_id(author_parse.author_name)
            author_ref = Ref("authors", author_id)
            totals.authors.add(author_id)
            # Shared entity: write only if absent, never clobber human data (M2).
            if not write_shared_entity(
                backend, "authors", author_id,
                build_author_doc(forename, surname, author_gender, now),
            ):
                totals.authors_existing += 1

        # Illustrator (shared entity — write only if absent, #73 M2/S1).
        illustrator_ref = None
        if not is_unknown(illustrator_name):
            ill_id = person_document_id(illustrator_name)
            illustrator_ref = Ref("illustrators", ill_id)
            totals.illustrators.add(ill_id)
            if not write_shared_entity(
                backend, "illustrators", ill_id, build_named_doc(illustrator_name, now)
            ):
                totals.illustrators_existing += 1

        # Publisher (shared entity — write only if absent, #73 M2/S1).
        publisher_ref = None
        if not is_unknown(publisher_name):
            pub_id = person_document_id(publisher_name)
            publisher_ref = Ref("publishers", pub_id)
            totals.publishers.add(pub_id)
            if not write_shared_entity(
                backend, "publishers", pub_id, build_named_doc(publisher_name, now)
            ):
                totals.publishers_existing += 1

        # Characters.
        character_refs: list = []
        char_index_to_docid: dict = {}
        seen_char_ids: set = set()
        char_plans: list = []
        for row in db_chars:
            char_id = character_document_id(book_id, row["name"])
            if char_id in seen_char_ids:
                continue
            seen_char_ids.add(char_id)
            human, _ = map_human(row["human"])
            gender = map_character_gender(row["gender"])
            backend.write_document(
                "characters",
                char_id,
                build_character_doc(
                    book_ref=book_ref,
                    name=row["name"],
                    gender=gender,
                    protagonist=bool(row["is_protagonist"]),
                    human=human,
                    now=now,
                ),
            )
            character_refs.append(Ref("characters", char_id))
            char_index_to_docid[row["index"]] = char_id
            totals.characters += 1
            char_plans.append(
                f"{row['name']} (g={gender}, human={human}, protag={bool(row['is_protagonist'])})"
            )

        # Aliases (link to the character resolved via character_id).
        alias_plans: list = []
        for arow in db_aliases:
            target_docid = char_index_to_docid.get(arow["character_id"])
            if target_docid is None:
                alias_plans.append(
                    f"{arow['alias']!r} -> UNRESOLVED character_id={arow['character_id']}"
                )
                continue
            alias_id = alias_document_id(book_id, arow["alias"])
            backend.write_document(
                "aliases",
                alias_id,
                build_alias_doc(
                    character_ref=Ref("characters", target_docid),
                    book_ref=book_ref,
                    name=arow["alias"],
                    now=now,
                ),
            )
            totals.aliases += 1
            alias_plans.append(f"{arow['alias']!r} -> {arow['character']!r}")

        # Pages (render every physical PDF page; text for story pages).
        page_count = 0
        story_page_count = 0
        pdf_analysis = None
        sample_page_text = None
        # Concatenated final story-page text, used to cross-check against the
        # validated (pickle) reference for this book.
        book_story_text_parts: list = []
        # Page numbers (and reasons) the clean pass flagged as incoherent.
        review_pages: list = []
        review_notes: list = []
        if pdf_path:
            pdf_analysis = analyse_pdf(pdf_path)
            page_count = pdf_analysis["page_count"]
            page_texts = pdf_analysis["page_texts"]
            totals.pages_text_layer += pdf_analysis["text_layer_pages"]

            if args.execute:
                import pypdfium2 as pdfium

                # Pass 1: render every page to S3 and extract its RAW text
                # (text layer if present, else Sonnet-5 OCR — unless a wordless
                # spread flanked by text-layer pages lets us skip OCR entirely).
                # We collect all pages first so pass 2 can judge each story page
                # in the context of its neighbours.
                #
                # Neighbour-continuity skip (DECISIONS 010): a story page with no
                # text layer, sitting between two TEXT-LAYER story pages, is a
                # candidate to skip OCR if the story reads continuously across it
                # (a genuine wordless spread). ``story_is_text_layer`` marks which
                # story pages have a real PDF text layer (never OCR'd), so the
                # judge only ever sees true text-layer neighbours, as validated.
                story_indices = [
                    idx for idx in range(page_count)
                    if is_story_page(idx + 1, book.page_range)
                ]
                story_pos_of = {idx: pos for pos, idx in enumerate(story_indices)}
                story_is_text_layer = [
                    len(page_texts[idx] if idx < len(page_texts) else "") >= TEXT_LAYER_MIN_CHARS
                    for idx in story_indices
                ]
                page_rows: list = []
                pdf_doc = pdfium.PdfDocument(pdf_path)
                try:
                    for i in range(page_count):
                        page_number = i + 1
                        in_story = is_story_page(page_number, book.page_range)
                        jpeg = render_page_jpeg(pdf_doc, i, args.max_edge, args.jpeg_quality)
                        backend.s3_put(s3_page_path(book.title, page_number), jpeg)
                        totals.images += 1

                        raw = ""
                        text_source = "none"
                        ocr_error: Optional[str] = None
                        ocr_model_used = ""
                        continuity_verdict: Optional[dict] = None
                        if in_story:
                            layer = page_texts[i] if i < len(page_texts) else ""
                            if len(layer) >= TEXT_LAYER_MIN_CHARS:
                                raw = layer
                                text_source = "layer"
                            elif ai_client is not None:
                                # Image-only story page: try the neighbour-continuity
                                # skip before paying for OCR. Only the flanked case
                                # (interior page with text-layer neighbours on BOTH
                                # sides) qualifies; edges, runs of image-only pages,
                                # and any image-only neighbour → OCR unconditionally.
                                pos = story_pos_of.get(i, -1)
                                if pos >= 0 and flanked_by_text_layer(
                                    pos, len(story_indices), story_is_text_layer
                                ):
                                    prev_layer = page_texts[story_indices[pos - 1]]
                                    next_layer = page_texts[story_indices[pos + 1]]
                                    cont_key = (
                                        "continuity:"
                                        + sha1_hex(prev_layer + chr(0) + next_layer)
                                    )
                                    verdict = cache.get(cont_key)
                                    if verdict is None:
                                        # Fail-safe: check_narrative_continuity degrades
                                        # to an OCR-forcing verdict on any error, so a
                                        # judge failure never skips and never counts
                                        # toward the OCR circuit-breaker.
                                        verdict = check_narrative_continuity(
                                            ai_client, prev_layer, next_layer,
                                            model=args.continuity_model,
                                        )
                                        cache.put(cont_key, verdict)
                                    if should_skip_ocr(verdict):
                                        continuity_verdict = verdict
                                if continuity_verdict is not None:
                                    # Genuine wordless spread — store empty text, do
                                    # not OCR, but record the verdict for auditability.
                                    text_source = "skipped_wordless"
                                    totals.pages_skipped_ocr += 1
                                else:
                                    text_source = "ocr"
                                    ocr_model_used = args.ocr_model
                                    totals.pages_needing_ocr += 1
                                    ocr_key = f"ocr:{book_id}:{page_number}:{sha1_hex(jpeg)}"
                                    cached = cache.get(ocr_key)
                                    if cached is not None:
                                        raw = str(cached.get("text") or "")
                                    else:
                                        raw, ocr_error = ai_ocr_page(
                                            ai_client, jpeg, args.ocr_model
                                        )
                                        if ocr_error is None:
                                            consecutive_ai_failures = 0
                                            cache.put(ocr_key, {"text": raw})
                                        else:
                                            totals.ocr_failed += 1
                                            _note_ai_failure()
                            # Quote/tag/punctuation-only output means a blank page.
                            raw = normalise_blank_text(raw)
                        page_rows.append({
                            "page_number": page_number,
                            "in_story": in_story,
                            "raw": raw,
                            "text_source": text_source,
                            "ocr_error": ocr_error,
                            "ocr_model": ocr_model_used,
                            "continuity": continuity_verdict,
                        })
                finally:
                    pdf_doc.close()

                # Pass 2: clean + judge each story page in context of its
                # neighbouring story pages, then persist every page.
                story_positions = [k for k, r in enumerate(page_rows) if r["in_story"]]
                last_sp = len(story_positions) - 1
                for sp, k in enumerate(story_positions):
                    row = page_rows[k]
                    text = row["raw"]
                    status = ""
                    row["clean_model"] = ""

                    if row["ocr_error"]:
                        # OCR call failed (not a genuine blank) — flag high (M1).
                        row["text"] = ""
                        row["needs_review"] = True
                        row["priority"] = "high"
                        row["note"] = row["ocr_error"]
                    elif not args.no_clean and ai_client is not None:
                        # Position-derived neighbour context (M4): distinguishes a
                        # true book edge from an interior wordless/missed page.
                        prev_raw = page_rows[story_positions[sp - 1]]["raw"] if sp > 0 else ""
                        next_raw = (
                            page_rows[story_positions[sp + 1]]["raw"] if sp < last_sp else ""
                        )
                        prev_ctx = neighbour_context(
                            prev_raw, is_previous=True, at_book_edge=(sp == 0)
                        )
                        next_ctx = neighbour_context(
                            next_raw, is_previous=False, at_book_edge=(sp == last_sp)
                        )
                        clean_key = (
                            f"clean:{book_id}:{row['page_number']}:"
                            f"{sha1_hex(text + chr(0) + prev_ctx + chr(0) + next_ctx)}"
                        )
                        cached = cache.get(clean_key)
                        if cached is not None:
                            text = cached.get("text", text)
                            status = cached.get("status", "unchanged")
                            needs_review = bool(cached.get("needs_review", False))
                            priority = cached.get("priority", "")
                            note = cached.get("note", "")
                        else:
                            text, status, needs_review, priority, note = ai_clean_and_judge(
                                ai_client, row["raw"], prev_ctx, next_ctx, args.clean_model
                            )
                            if status == "failed":
                                totals.judge_failed += 1
                                _note_ai_failure()
                            else:
                                consecutive_ai_failures = 0
                                cache.put(clean_key, {
                                    "text": text, "status": status,
                                    "needs_review": needs_review,
                                    "priority": priority, "note": note,
                                })
                        row["clean_model"] = args.clean_model
                        if status == "cleaned":
                            totals.pages_cleaned += 1
                        elif status == "rejected":
                            totals.pages_clean_rejected += 1
                        row["text"] = text
                        row["needs_review"] = needs_review
                        row["priority"] = priority
                        row["note"] = note
                    else:
                        row["text"] = text
                        row["needs_review"] = False
                        row["priority"] = ""
                        row["note"] = ""

                    row["clean_status"] = status
                    if row["text"]:
                        book_story_text_parts.append(row["text"])
                    if row["needs_review"]:
                        review_pages.append(row["page_number"])
                        review_notes.append((row["page_number"], row["priority"], row["note"]))
                        totals.pages_flagged += 1
                        if row["priority"] == "high":
                            totals.pages_flagged_high += 1

                for row in page_rows:
                    if row["in_story"]:
                        story_page_count += 1
                    backend.write_document(
                        "pages",
                        page_document_id(book_id, row["page_number"]),
                        build_page_doc(
                            book_ref=book_ref,
                            page_number=row["page_number"],
                            contains_story=row["in_story"],
                            text=row.get("text", "") if row["in_story"] else "",
                            now=now,
                            needs_review=row.get("needs_review", False),
                            review_note=row.get("note", ""),
                            review_priority=row.get("priority", ""),
                            text_source=row.get("text_source", "none"),
                            clean_status=row.get("clean_status", ""),
                            ocr_model=row.get("ocr_model", ""),
                            clean_model=row.get("clean_model", ""),
                            continuity=row.get("continuity"),
                        ),
                    )
                    totals.pages += 1
            else:
                # Dry-run accounting: count planned pages / story pages / OCR
                # needs without rendering or writing. The cross-check uses the
                # text-layer text as a preview of extraction quality (OCR and
                # cleaning only run on --execute).
                for i in range(page_count):
                    if is_story_page(i + 1, book.page_range):
                        story_page_count += 1
                        layer = page_texts[i] if i < len(page_texts) else ""
                        if len(layer) < TEXT_LAYER_MIN_CHARS:
                            totals.pages_needing_ocr += 1
                        elif layer:
                            book_story_text_parts.append(layer)
                totals.pages += page_count
                totals.images += page_count
                for i, txt in enumerate(page_texts):
                    if is_story_page(i + 1, book.page_range) and txt:
                        sample_page_text = (i + 1, txt[:120])
                        break

        totals.story_pages += story_page_count
        totals.books += 1

        # --- Cross-check extracted text vs the validated reference -----------
        # Primary metric is CONTAINMENT (recall of the validated text); Jaccard
        # is kept as a secondary "excess text" indicator (#73 S3).
        reference_text = validated_text.get(book.norm, "")
        containment = None
        jaccard = None
        if reference_text:
            extracted = "\n".join(book_story_text_parts)
            containment = text_containment(extracted, reference_text)
            jaccard = text_similarity(extracted, reference_text)
            totals.similarities.append((book.title, containment, jaccard))
        else:
            totals.no_reference.append(book.title)

        high_priority = any(pri == "high" for _, pri, _ in review_notes)
        if review_pages:
            totals.books_flagged.append((book.title, list(review_pages), high_priority))

        # --- Book-level review reasons (#73 M3): empty / character-less books --
        photos_uploaded = page_count > 0
        if not pdf_path or page_count == 0:
            book_review_notes.append("no PDF / page images for this book")
        if not character_refs:
            book_review_notes.append("no characters recorded for this book")
        book_review_note = "; ".join(book_review_notes)
        book_needs_review = bool(review_pages) or bool(book_review_notes)

        # --- Book document (the record every page/character references) -------
        backend.write_document(
            "books",
            book_id,
            build_book_doc(
                title=book.title,
                author_ref=author_ref,
                illustrator_ref=illustrator_ref,
                publisher_ref=publisher_ref,
                published=published,
                page_range=book.page_range,
                page_count=page_count,
                character_refs=character_refs,
                photos_uploaded=photos_uploaded,
                needs_review=book_needs_review,
                review_pages=review_pages,
                high_priority_review=high_priority,
                review_note=book_review_note,
                now=now,
            ),
        )

        # --- Per-book plan line ------------------------------------------
        author_display = author_parse.author_name or "(none — flagged)"
        second = f"  (+2nd: {author_parse.second_name})" if author_parse.second_name else ""
        print(f"BOOK: {book.title}  [{book_id}]")
        print(f"    author      : {author_display} -> gender {author_gender}{second}")
        print(f"    illustrator : {illustrator_name}{'  (AI)' if do_lookup else ''}")
        print(f"    publisher   : {publisher_name}{'  (AI)' if do_lookup else ''}")
        print(f"    published   : {published if published is not None else '(unknown)'}")
        print(f"    page range  : {book.page_range or '(all pages)'}")
        print(f"    pdf         : {os.path.basename(pdf_path) if pdf_path else 'NONE (no images/pages)'}")
        if pdf_analysis:
            print(f"    pages       : {page_count} total, {story_page_count} story, "
                  f"text-layer {pdf_analysis['text_layer_pages']} / "
                  f"image-only {pdf_analysis['image_only_pages']}")
        print(f"    characters  : {len(character_refs)}")
        for line in char_plans:
            print(f"        - {line}")
        print(f"    aliases     : {len(alias_plans)}")
        for line in alias_plans:
            print(f"        - {line}")
        if containment is not None:
            kind = "text-layer" if not args.execute else "final"
            flag = "  <-- LOW, review" if containment < args.containment_threshold else ""
            print(f"    text match  : containment {containment:.0%} / jaccard "
                  f"{jaccard:.0%} ({kind} vs validated){flag}")
        elif validated_source != "none":
            print("    text match  : (no validated reference for this title)")
        for reason in book_review_notes:
            print(f"    review      : {reason}")
        if review_pages:
            print(f"    flagged     : {len(review_pages)} page(s) need review "
                  f"(re-extract / human read):")
            for pnum, pri, note in review_notes:
                tag = "HIGH" if pri == "high" else "normal"
                print(f"        - page {pnum} [{tag}]: {note or '(no reason given)'}")
        if sample_page_text and not args.execute:
            print(f"    sample text : page {sample_page_text[0]}: {sample_page_text[1]!r}")
        print()
    except AICircuitBreakerError as exc:
        aborted = True
        print()
        print("=" * 78)
        print(f"RUN ABORTED (circuit breaker): {exc}")
        print("=" * 78)
        print()

    # --- Corpus totals ----------------------------------------------------
    print("=" * 78)
    print("TOTALS")
    print("=" * 78)
    print(f"  books to import   : {totals.books}")
    print(f"  books skipped     : {totals.books_skipped} (already in Firestore)")
    print(f"  authors           : {len(totals.authors)}")
    print(f"  illustrators      : {len(totals.illustrators)}")
    print(f"  publishers        : {len(totals.publishers)}")
    print(f"  pages             : {totals.pages}  ({totals.story_pages} story)")
    print(f"  page images (S3)  : {totals.images}")
    print(f"  characters        : {totals.characters}")
    print(f"  aliases           : {totals.aliases}")
    print(f"  story pages w/ text layer   : {totals.pages_text_layer}")
    print(f"  story pages needing AI OCR  : {totals.pages_needing_ocr}")
    # OCR calls avoided by the neighbour-continuity skip (execute-only; the skip
    # requires the AI judge so it never fires in a dry run). Denominator is the
    # image-only story pages that were candidates for OCR (OCR'd + skipped).
    _ocr_candidates = totals.pages_needing_ocr + totals.pages_skipped_ocr
    _skip_pct = (totals.pages_skipped_ocr / _ocr_candidates * 100) if _ocr_candidates else 0.0
    print(f"  story pages OCR SKIPPED (wordless, neighbour-continuity) : "
          f"{totals.pages_skipped_ocr}  "
          f"({_skip_pct:.0f}% of image-only story pages — OCR calls avoided)")
    print(f"  shared-entity writes skipped (already existed, M2) : "
          f"authors {totals.authors_existing}, illustrators {totals.illustrators_existing}, "
          f"publishers {totals.publishers_existing}")
    if args.execute:
        print(f"  story pages junk-cleaned    : {totals.pages_cleaned}")
        print(f"  clean rewrites rejected by guard (kept original) : {totals.pages_clean_rejected}")
        print(f"  pages flagged needs_review  : {totals.pages_flagged} "
              f"({totals.pages_flagged_high} high-priority) across "
              f"{len(totals.books_flagged)} books")
        print(f"  OCR calls failed            : {totals.ocr_failed}")
        print(f"  clean/judge calls failed    : {totals.judge_failed}")
    print()

    if args.execute and totals.books_flagged:
        print("FLAGGED FOR REVIEW (fails coherence and/or context-fit)")
        print("  [HIGH] = fails both checks (strongest re-extract / human-read signal)")
        for title, pages, high in totals.books_flagged:
            tag = " [HIGH]" if high else ""
            print(f"  - {title}{tag}: pages {', '.join(str(p) for p in pages)}")
        print()

    # --- Text cross-check summary ----------------------------------------
    print("TEXT CROSS-CHECK (extracted vs validated reference)")
    print("  primary metric = CONTAINMENT (recall of the validated text); "
          "jaccard is a secondary 'excess text' indicator (#73 S3)")
    kind = "text-layer preview" if not args.execute else "final imported text"
    if totals.similarities:
        containments = [c for _, c, _ in totals.similarities]
        jaccards = [j for _, _, j in totals.similarities]
        avg_c = sum(containments) / len(containments)
        avg_j = sum(jaccards) / len(jaccards)
        low = sorted(
            [(t, c, j) for t, c, j in totals.similarities if c < args.containment_threshold],
            key=lambda x: x[1],
        )
        print(f"  compared ({kind}) : {len(containments)} books")
        print(f"  mean containment   : {avg_c:.0%}")
        print(f"  mean jaccard       : {avg_j:.0%}")
        print(f"  books below {args.containment_threshold:.0%} containment ({len(low)}):")
        for title, c, j in low:
            print(f"      - containment {c:.0%} / jaccard {j:.0%}  {title}")
    else:
        print("  (no books had a validated-text reference to compare against)")
    if totals.no_reference:
        print(f"  books with NO validated reference ({len(totals.no_reference)}):")
        for title in sorted(totals.no_reference):
            print(f"      - {title}")
    print()
    if aborted:
        print("RUN ABORTED by the AI circuit breaker — re-run to resume "
              "(already-written books are skipped; the aborted book is re-done).")
        return 1
    if not args.execute:
        print("DRY RUN complete — nothing was written to Firestore or S3.")
        print("Re-run with --execute (where the real secrets exist) to import.")
    else:
        print("EXECUTE complete — records written to Firestore + S3.")
    return 0


# ---------------------------------------------------------------------------
# CLI.
# ---------------------------------------------------------------------------

def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Import the fair-tales pilot corpus into Firestore + S3 (dry-run by default).",
    )
    p.add_argument(
        "--methods-dir",
        default="../fair-tales-language-analysis",
        help="Path to the pilot-corpus checkout holding Book-List-Final-NONA.xlsx, "
        "character_database.db, text_pdfs/ and data/ (default: "
        "../fair-tales-language-analysis).",
    )
    p.add_argument("--excel", help="Override path to Book-List-Final-NONA.xlsx.")
    p.add_argument("--db", help="Override path to character_database.db.")
    p.add_argument("--pdf-dir", help="Override path to the text_pdfs directory.")
    p.add_argument(
        "--pickle",
        help="Override path to data/book_dataframe.pickle (the human-validated "
        "text used to cross-check extraction).",
    )
    p.add_argument(
        "--json",
        help="Override path to data/book_dataframe.json (validated-text fallback "
        "if the pickle cannot be read).",
    )
    p.add_argument(
        "--secrets",
        default=DEFAULT_SECRETS,
        help=f"Path to .streamlit/secrets.toml (default: {DEFAULT_SECRETS}).",
    )
    p.add_argument(
        "--execute",
        action="store_true",
        help="Perform the real import (writes to Firestore + S3). Omit for a dry run.",
    )
    p.add_argument(
        "--limit", type=int, default=0, help="Only process the first N matched books (0 = all)."
    )
    p.add_argument(
        "--overwrite",
        action="store_true",
        help="Re-import books that already exist in Firestore (default skips them). "
        "Use to re-run after a fix; writes use set(merge=True).",
    )
    p.add_argument(
        "--sample-ai",
        type=int,
        default=1,
        help="In a dry run, run this many real AI illustrator/publisher lookups as a sample (default 1; 0 disables).",
    )
    p.add_argument("--lookup-model", default=DEFAULT_LOOKUP_MODEL, help="Claude model for the book-metadata web lookup.")
    p.add_argument("--ocr-model", default=DEFAULT_OCR_MODEL, help="Claude model for per-page OCR fallback.")
    p.add_argument(
        "--continuity-model",
        default=DEFAULT_CONTINUITY_MODEL,
        help=f"Claude model for the text-only neighbour-continuity judge that skips "
        f"OCR on wordless spreads flanked by text-layer pages "
        f"(default: {DEFAULT_CONTINUITY_MODEL}).",
    )
    p.add_argument(
        "--clean-model",
        default=DEFAULT_CLEAN_MODEL,
        help=f"Claude model for the per-page junk-character clean-up pass "
        f"(default: {DEFAULT_CLEAN_MODEL}).",
    )
    p.add_argument(
        "--no-clean",
        action="store_true",
        help="Disable the AI junk-character clean-up pass (import raw extracted text).",
    )
    p.add_argument(
        "--compare-threshold",
        type=float,
        default=DEFAULT_COMPARE_THRESHOLD,
        help=f"Secondary Jaccard word-overlap threshold, reported only "
        f"(default: {DEFAULT_COMPARE_THRESHOLD}).",
    )
    p.add_argument(
        "--containment-threshold",
        type=float,
        default=DEFAULT_CONTAINMENT_THRESHOLD,
        help=f"PRIMARY cross-check: flag books whose extracted text captures less "
        f"than this fraction of the validated reference's words "
        f"(default: {DEFAULT_CONTAINMENT_THRESHOLD}).",
    )
    p.add_argument(
        "--cache-dir",
        default=DEFAULT_CACHE_DIR,
        help=f"Local result cache for OCR/clean+judge calls, so a crash or "
        f"--overwrite re-run reuses prior results (default: {DEFAULT_CACHE_DIR}).",
    )
    p.add_argument(
        "--no-cache",
        action="store_true",
        help="Disable the local result cache (always call the API).",
    )
    p.add_argument("--max-edge", type=int, default=DEFAULT_MAX_EDGE, help="Longest-edge px when rendering page JPEGs.")
    p.add_argument("--jpeg-quality", type=int, default=DEFAULT_JPEG_QUALITY, help="JPEG quality for page images.")
    return p


def main(argv: Optional[list] = None) -> int:
    args = build_arg_parser().parse_args(argv)
    return plan_and_run(args)


if __name__ == "__main__":
    raise SystemExit(main())
